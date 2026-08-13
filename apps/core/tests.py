from datetime import date
from pathlib import Path
from unittest.mock import Mock, patch

from django.conf import settings
from django.core import mail
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from apps.accounts.models import GoogleCalendarConnection, PushSubscription, User
from apps.articles.models import Article
from config.storage import supabase_public_domain

from . import google_calendar as google_calendar_module
from . import push as push_module
from .models import SESSION_THEME_KEY, Announcement, ContactLink, SiteConfig, Theme, get_effective_theme
from .notifications import notifications_feed, unread_notifications_count


class HomeTests(TestCase):
    def test_reparte_los_articulos_mas_recientes_entre_destacados_y_ultimos(self):
        for i in range(10):
            Article.objects.create(title=f"Artículo {i}", body="Cuerpo")
        response = self.client.get(reverse("core:home"))
        self.assertEqual(len(response.context["hero_articles"]), 3)
        self.assertEqual(len(response.context["grid_articles"]), 5)
        # Sin solapamiento: ninguno de los "últimos artículos" repite uno
        # que ya salga en el carrusel destacado.
        hero_pks = {a.pk for a in response.context["hero_articles"]}
        grid_pks = {a.pk for a in response.context["grid_articles"]}
        self.assertEqual(hero_pks & grid_pks, set())

    def test_el_titulo_enlaza_al_articulo(self):
        article = Article.objects.create(title="Mi artículo de prueba", body="Cuerpo")
        response = self.client.get(reverse("core:home"))
        self.assertContains(response, article.get_absolute_url())
        self.assertContains(response, "Mi artículo de prueba")

    def test_sin_articulos_muestra_el_hero_generico(self):
        response = self.client.get(reverse("core:home"))
        self.assertEqual(list(response.context["hero_articles"]), [])
        self.assertContains(response, "Bienvenido a")

    def test_invitado_ve_el_banner_de_registro_junto_al_carrusel(self):
        Article.objects.create(title="Artículo con carrusel", body="Cuerpo")
        response = self.client.get(reverse("core:home"))
        self.assertContains(response, "home-guest-banner")
        self.assertContains(response, reverse("accounts:register"))

    def test_usuario_logueado_no_ve_el_banner_de_registro(self):
        Article.objects.create(title="Artículo con carrusel", body="Cuerpo")
        user = User.objects.create(email="home_user@test.local", role=User.Role.LECTOR)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=user.email, password="Testpass123!")
        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, "home-guest-banner")

    def test_la_fila_explora_enlaza_a_las_secciones_principales(self):
        response = self.client.get(reverse("core:home"))
        for url_name in ["articles:list", "forum:list", "movies:list", "games:hub", "shop:list", "secret:gate"]:
            self.assertContains(response, reverse(url_name))


class WrapAfterPeriodFilterTests(TestCase):
    def test_solo_deja_partible_el_espacio_tras_el_primer_punto(self):
        from apps.core.templatetags.core_extras import wrap_after_period

        result = wrap_after_period("Cine, pelis, series y debates. Todo en una misma sala.")
        head, sep, tail = result.partition(". ")
        self.assertEqual(sep, ". ")
        self.assertNotIn(" ", head)
        self.assertNotIn(" ", tail)

    def test_sin_punto_no_deja_ningun_espacio_partible(self):
        from apps.core.templatetags.core_extras import wrap_after_period

        result = wrap_after_period("Sin punto en la frase")
        self.assertNotIn(" ", result)

    def test_escapa_html(self):
        from apps.core.templatetags.core_extras import wrap_after_period

        result = wrap_after_period("<script>alert(1)</script>")
        self.assertNotIn("<script>", result)


class RecentActivityTests(TestCase):
    """Actividad reciente de la portada: últimos comentarios de Artículos
    y del Foro, mezclados por fecha, desactivable desde admin."""

    def setUp(self):
        self.user = _make_user("actividad@test.local")

    def test_un_comentario_de_articulo_sale_en_la_portada(self):
        from apps.articles.models import Article, ArticleComment

        article = Article.objects.create(title="Crítica de prueba", body="Cuerpo")
        ArticleComment.objects.create(article=article, author=self.user, body="Muy buena reseña.")

        response = self.client.get(reverse("core:home"))
        self.assertContains(response, "Muy buena reseña.")
        self.assertContains(response, f"{article.get_absolute_url()}#comment-")

    def test_un_comentario_de_foro_sale_en_la_portada(self):
        from apps.forum.models import Thread, ThreadComment

        thread = Thread.objects.create(title="Hilo de prueba", body="Cuerpo", author=self.user)
        ThreadComment.objects.create(thread=thread, author=self.user, body="Totalmente de acuerdo.")

        response = self.client.get(reverse("core:home"))
        self.assertContains(response, "Totalmente de acuerdo.")

    def test_comentario_de_articulo_privado_no_sale_para_quien_no_puede_verlo(self):
        from apps.articles.models import Article, ArticleComment

        article = Article.objects.create(title="Solo para el equipo", body="Cuerpo", is_private=True)
        ArticleComment.objects.create(article=article, author=self.user, body="Comentario privado.")

        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, "Comentario privado.")

    def test_comentario_borrado_del_foro_no_sale(self):
        from apps.forum.models import Thread, ThreadComment

        thread = Thread.objects.create(title="Hilo con borrado", body="Cuerpo", author=self.user)
        ThreadComment.objects.create(thread=thread, author=self.user, body="Esto se borra", is_deleted=True)

        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, "Esto se borra")

    def test_desactivada_desde_admin_no_sale_nada(self):
        from apps.forum.models import Thread, ThreadComment

        thread = Thread.objects.create(title="Hilo visible", body="Cuerpo", author=self.user)
        ThreadComment.objects.create(thread=thread, author=self.user, body="Comentario visible")

        config = SiteConfig.load()
        config.recent_activity_enabled = False
        config.save()

        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, "Comentario visible")
        self.assertNotContains(response, "Actividad reciente")


class ContactFormTests(TestCase):
    """"Escríbenos" ya no manda un email (el SMTP no era de fiar) — en su
    lugar, cada Admin recibe el mensaje como un aviso de Social."""

    def setUp(self):
        self.admin = User.objects.create(email="admin_contacto@test.local", role=User.Role.ADMIN)
        self.admin.set_password("Testpass123!")
        self.admin.save()

    def test_envio_valido_llega_como_mensaje_de_social_al_admin(self):
        from apps.social.models import Message

        response = self.client.post(reverse("core:contact"), {
            "name": "Ana",
            "email": "ana@example.com",
            "message": "Hola, os escribo para...",
            "website": "",
        })
        self.assertRedirects(response, reverse("core:contact"))
        msg = Message.objects.get(recipient=self.admin)
        self.assertEqual(msg.sender, self.admin)
        self.assertIn("Ana", msg.body)
        self.assertIn("ana@example.com", msg.body)
        self.assertIn("Hola, os escribo para...", msg.body)

    def test_llega_a_todos_los_admin_si_hay_varios(self):
        from apps.social.models import Message

        otro_admin = User.objects.create(email="admin2_contacto@test.local", role=User.Role.ADMIN)
        self.client.post(reverse("core:contact"), {
            "name": "Ana", "email": "ana@example.com", "message": "Hola", "website": "",
        })
        self.assertTrue(Message.objects.filter(recipient=self.admin).exists())
        self.assertTrue(Message.objects.filter(recipient=otro_admin).exists())

    def test_honeypot_relleno_no_manda_nada(self):
        from apps.social.models import Message

        response = self.client.post(reverse("core:contact"), {
            "name": "Bot",
            "email": "bot@example.com",
            "message": "comprar barato ahora",
            "website": "http://spam.example.com",
        })
        self.assertRedirects(response, reverse("core:contact"))
        self.assertEqual(Message.objects.count(), 0)


class ContactLinkTests(TestCase):
    """Enlaces de contacto alternativos (Instagram, WhatsApp...) en /contacto/."""

    def test_se_muestran_los_enlaces(self):
        ContactLink.objects.create(
            platform=ContactLink.Platform.INSTAGRAM, label="@lasaladebygui", url="https://instagram.com/lasaladebygui",
        )
        response = self.client.get(reverse("core:contact"))
        self.assertContains(response, "@lasaladebygui")
        self.assertContains(response, "https://instagram.com/lasaladebygui")

    def test_orden_respeta_el_campo_order(self):
        segundo = ContactLink.objects.create(platform=ContactLink.Platform.OTRO, label="Segundo", url="https://b.example.com", order=2)
        primero = ContactLink.objects.create(platform=ContactLink.Platform.OTRO, label="Primero", url="https://a.example.com", order=1)
        response = self.client.get(reverse("core:contact"))
        self.assertEqual(list(response.context["contact_links"]), [primero, segundo])

    def test_icono_por_defecto_para_plataforma_no_reconocida(self):
        link = ContactLink.objects.create(platform="algo-raro", label="X", url="https://example.com")
        self.assertEqual(link.icon, "🔗")

    def test_icono_de_una_plataforma_conocida(self):
        link = ContactLink(platform=ContactLink.Platform.INSTAGRAM)
        self.assertEqual(link.icon, "📷")


class DonationsPageTests(TestCase):
    def test_muestra_el_numero_de_bizum_configurado(self):
        config = SiteConfig.load()
        config.bizum_number = "600 000 000"
        config.save()
        response = self.client.get(reverse("core:donations"))
        self.assertContains(response, "600 000 000")


class IntroAnimationTests(TestCase):
    def test_se_muestra_por_defecto(self):
        response = self.client.get(reverse("core:home"))
        self.assertContains(response, 'id="intro"')

    def test_se_oculta_si_se_desactiva_desde_el_admin(self):
        config = SiteConfig.load()
        config.show_intro_animation = False
        config.save()
        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, 'id="intro"')

    def test_sin_sonido_personalizado_el_dato_va_vacio(self):
        response = self.client.get(reverse("core:home"))
        self.assertContains(response, "introOverlay('')")

    def test_con_sonido_personalizado_se_pasa_su_url(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        config = SiteConfig.load()
        config.intro_sound = SimpleUploadedFile("carrete.mp3", b"fake-audio-data", content_type="audio/mpeg")
        config.save()
        response = self.client.get(reverse("core:home"))
        self.assertContains(response, config.intro_sound.url)


class BootstrapProductionTests(TestCase):
    def test_sin_variables_no_hace_nada(self):
        call_command("bootstrap_production")
        self.assertFalse(User.objects.exists())

    def test_crea_admin_si_se_piden_las_variables(self):
        import os

        os.environ["DJANGO_SUPERUSER_EMAIL"] = "admin@lasaladebygui.local"
        os.environ["DJANGO_SUPERUSER_PASSWORD"] = "Testpass123!"
        try:
            call_command("bootstrap_production")
        finally:
            del os.environ["DJANGO_SUPERUSER_EMAIL"]
            del os.environ["DJANGO_SUPERUSER_PASSWORD"]

        user = User.objects.get(email="admin@lasaladebygui.local")
        self.assertEqual(user.role, User.Role.ADMIN)
        self.assertTrue(user.is_superuser)
        self.assertTrue(user.check_password("Testpass123!"))

    def test_no_crea_un_segundo_admin_si_ya_existe_uno(self):
        import os

        existing = User(email="ya-existe@lasaladebygui.local", role=User.Role.ADMIN)
        existing.set_password("Otra1234!")
        existing.save()

        os.environ["DJANGO_SUPERUSER_EMAIL"] = "otro@lasaladebygui.local"
        os.environ["DJANGO_SUPERUSER_PASSWORD"] = "Testpass123!"
        try:
            call_command("bootstrap_production")
        finally:
            del os.environ["DJANGO_SUPERUSER_EMAIL"]
            del os.environ["DJANGO_SUPERUSER_PASSWORD"]

        self.assertFalse(User.objects.filter(email="otro@lasaladebygui.local").exists())

    def test_run_seed_quotes_carga_las_frases(self):
        import os

        from apps.games.models import MovieQuote

        os.environ["RUN_SEED_QUOTES"] = "true"
        try:
            call_command("bootstrap_production")
        finally:
            del os.environ["RUN_SEED_QUOTES"]

        self.assertTrue(MovieQuote.objects.exists())

    def test_sin_run_seed_quotes_no_carga_nada(self):
        from apps.games.models import MovieQuote

        call_command("bootstrap_production")
        self.assertFalse(MovieQuote.objects.exists())

    @override_settings(TMDB_API_KEY="")
    def test_el_contenido_de_juegos_se_carga_siempre_sin_variables(self):
        """A diferencia de las películas y las frases (gated por variables
        de entorno, porque cuestan llamadas a APIs), Trivial/Emoji/Malas
        descripciones/Actor/Verdadero o falso, Oscar y Qué personaje eres
        se cargan siempre — son baratos e idempotentes."""
        from apps.games.models import OscarCategory, PersonalityCharacter, TriviaQuestion

        call_command("bootstrap_production")
        self.assertTrue(TriviaQuestion.objects.exists())
        self.assertTrue(OscarCategory.objects.exists())
        self.assertTrue(PersonalityCharacter.objects.exists())


class ServiceWorkerTests(TestCase):
    """El service worker se sirve en /sw.js (raíz), no bajo /static/js/: es
    lo que le da scope de todo el sitio en vez de solo /static/js/."""

    def test_se_sirve_en_la_raiz_con_scope_de_todo_el_sitio(self):
        response = self.client.get("/sw.js")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Service-Worker-Allowed"], "/")
        self.assertIn("javascript", response["Content-Type"])


class ThemeSwitcherTests(TestCase):
    def setUp(self):
        self.noir = Theme.objects.get(slug="noir")
        self.vintage = Theme.objects.get(slug="vintage")

    def test_anonimo_cambia_tema_via_sesion(self):
        response = self.client.post(reverse("core:set-theme", args=[self.noir.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.client.session[SESSION_THEME_KEY], "noir")

        theme = get_effective_theme(None, self.client.session)
        self.assertEqual(theme, self.noir)

    def test_anonimo_ve_el_tema_en_theme_css(self):
        self.client.post(reverse("core:set-theme", args=[self.vintage.slug]))
        response = self.client.get(reverse("theme-css"))
        self.assertContains(response, self.vintage.color_bg)

    def test_usuario_logueado_guarda_en_su_cuenta_no_en_sesion(self):
        user = User.objects.create(email="lector@test.local", role=User.Role.LECTOR)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=user.email, password="Testpass123!")

        self.client.post(reverse("core:set-theme", args=[self.noir.slug]))
        user.refresh_from_db()
        self.assertEqual(user.theme, self.noir)
        self.assertNotIn(SESSION_THEME_KEY, self.client.session)

    def test_reset_theme_quita_la_preferencia(self):
        user = User.objects.create(email="lector2@test.local", role=User.Role.LECTOR, theme=self.noir)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=user.email, password="Testpass123!")

        self.client.post(reverse("core:reset-theme"))
        user.refresh_from_db()
        self.assertIsNone(user.theme)

    def test_theme_css_no_se_cachea_publicamente(self):
        response = self.client.get(reverse("theme-css"))
        self.assertIn("no-cache", response.headers["Cache-Control"])
        self.assertIn("private", response.headers["Cache-Control"])

    def test_existe_el_tema_jinx(self):
        jinx = Theme.objects.get(slug="jinx")
        self.assertTrue(jinx.is_published)
        self.assertEqual(jinx.color_accent, "#00E5FF")
        self.assertEqual(jinx.color_accent_secondary, "#FF2FD0")

    def test_existe_el_tema_blanco_y_negro(self):
        theme = Theme.objects.get(slug="blanco-y-negro")
        self.assertTrue(theme.is_published)
        self.assertTrue(theme.is_dark)

    def test_set_theme_con_next_redirige_en_vez_de_devolver_json(self):
        response = self.client.post(
            reverse("core:set-theme", args=[self.noir.slug]), {"next": reverse("accounts:settings")},
        )
        self.assertRedirects(response, reverse("accounts:settings"), fetch_redirect_response=False)

    def test_set_theme_con_next_externo_no_redirige_ahi(self):
        response = self.client.post(
            reverse("core:set-theme", args=[self.noir.slug]), {"next": "https://evil.example/"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"ok": True, "slug": self.noir.slug})

    def test_reset_theme_con_next_redirige(self):
        response = self.client.post(reverse("core:reset-theme"), {"next": reverse("accounts:settings")})
        self.assertRedirects(response, reverse("accounts:settings"), fetch_redirect_response=False)

    def test_un_tema_despublicado_no_aparece_en_el_selector(self):
        self.noir.is_published = False
        self.noir.save(update_fields=["is_published"])

        response = self.client.get(reverse("core:home"))
        self.assertNotIn(self.noir, response.context["all_themes"])
        self.assertIn(self.vintage, response.context["all_themes"])

    def test_el_selector_respeta_el_orden_configurado(self):
        self.noir.order = 1
        self.noir.save(update_fields=["order"])
        self.vintage.order = 0
        self.vintage.save(update_fields=["order"])

        response = self.client.get(reverse("core:home"))
        themes = list(response.context["all_themes"])
        self.assertLess(themes.index(self.vintage), themes.index(self.noir))

    def test_un_tema_despublicado_sigue_funcionando_si_ya_estaba_puesto(self):
        user = User.objects.create(email="lector3@test.local", role=User.Role.LECTOR, theme=self.noir)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=user.email, password="Testpass123!")

        self.noir.is_published = False
        self.noir.save(update_fields=["is_published"])

        response = self.client.get(reverse("theme-css"))
        self.assertContains(response, self.noir.color_bg)


class SupabasePublicDomainTests(TestCase):
    """Marcar un bucket de Supabase como 'Public' no lo hace accesible por
    la URL del endpoint S3 (esa exige petición firmada siempre) — hace
    falta la URL nativa .../storage/v1/object/public/<bucket>/... Esto
    verifica que la construimos bien a partir del endpoint S3 configurado."""

    def test_deriva_el_dominio_publico_desde_el_endpoint_s3(self):
        domain = supabase_public_domain(
            "https://xpleukpsphqwuvshyyls.storage.supabase.co/storage/v1/s3", "media",
        )
        self.assertEqual(
            domain,
            "xpleukpsphqwuvshyyls.supabase.co/storage/v1/object/public/media",
        )

    def test_usa_el_nombre_del_bucket_configurado(self):
        domain = supabase_public_domain(
            "https://otro-proyecto.storage.supabase.co/storage/v1/s3", "avatares",
        )
        self.assertIn("/object/public/avatares", domain)


class AdminAccessTests(TestCase):
    """/admin/ debe ser solo para el Admin (is_superuser) — Gestor y Editor
    son is_staff (para sus permisos puntuales dentro del propio /admin/,
    como el Gestor con el foro) pero no deben poder entrar al panel."""

    def _login_as(self, email, role):
        user = User.objects.create(email=email, role=role)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=email, password="Testpass123!")
        return user

    def test_admin_entra(self):
        self._login_as("admin@test.local", User.Role.ADMIN)
        response = self.client.get(reverse("admin:index"))
        self.assertEqual(response.status_code, 200)

    def test_formulario_de_edicion_tiene_boton_cancelar(self):
        """El botón "Cancelar" (templates/admin/submit_line.html, sin
        condición de popup) debe verse en CUALQUIER formulario de edición
        del admin, no solo en el de Theme — por eso se prueba aquí con
        Theme, pero el override es genérico para toda la app."""
        self._login_as("admin_cancel@test.local", User.Role.ADMIN)
        theme = Theme.objects.first()
        response = self.client.get(reverse("admin:core_theme_change", args=[theme.pk]))
        self.assertContains(response, "Cancelar")

    def test_gestor_no_entra(self):
        self._login_as("gestor@test.local", User.Role.GESTOR)
        response = self.client.get(reverse("admin:index"))
        self.assertRedirects(response, "/admin/login/?next=/admin/", fetch_redirect_response=False)

    def test_editor_no_entra(self):
        self._login_as("editor@test.local", User.Role.EDITOR)
        response = self.client.get(reverse("admin:index"))
        self.assertRedirects(response, "/admin/login/?next=/admin/", fetch_redirect_response=False)

    def test_anonimo_no_entra(self):
        response = self.client.get(reverse("admin:index"))
        self.assertRedirects(response, "/admin/login/?next=/admin/", fetch_redirect_response=False)


class ThemeAdminFormTests(TestCase):
    """El admin de Temas usa un desplegable para las tipografías (solo las
    que están cargadas de verdad en el sitio) y ya no deja tocar el ancho
    máximo de contenido desde el formulario."""

    def setUp(self):
        user = User.objects.create(email="theme_admin@test.local", role=User.Role.ADMIN, is_staff=True, is_superuser=True)
        user.set_password("Testpass123!")
        user.save()
        self.client.login(username=user.email, password="Testpass123!")
        self.theme = Theme.objects.get(slug="cinephile")

    def test_tipografia_es_un_desplegable(self):
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertContains(response, '<select name="font_heading"')
        self.assertContains(response, "Bebas Neue (grande, tipo cartel)")

    def test_hay_mas_tipografias_disponibles(self):
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertContains(response, "Cinzel (clásica, tipo épica/romana)")
        self.assertContains(response, "Oswald (condensada, titulares)")
        self.assertContains(response, "Cormorant Garamond (fina, romántica)")
        self.assertContains(response, "Poppins (geométrica, redondeada)")

    def test_no_se_puede_editar_el_ancho_maximo_desde_el_formulario(self):
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertNotContains(response, 'name="max_content_width"')

    def test_los_colores_de_la_animacion_de_inicio_son_editables(self):
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertContains(response, 'name="color_intro_light"')
        self.assertContains(response, 'name="color_intro_lamp"')
        self.assertContains(response, 'name="color_intro_chair"')
        self.assertContains(response, 'type="color"')

    def test_los_colores_estan_todos_juntos_en_un_unico_apartado(self):
        # Antes había 5 apartados de colores por separado (Fondo y
        # superficies, Texto, Acento principal...); ahora están todos juntos
        # en "Colores", en cuadrícula — menos compartimentado.
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertContains(response, "theme-color-grid")
        self.assertNotContains(response, "Fondo y superficies")

    def test_ya_no_hay_vista_previa_por_iframe(self):
        response = self.client.get(reverse("admin:core_theme_change", args=[self.theme.pk]))
        self.assertNotContains(response, "theme-preview-frame")


class SortableAdminTests(TestCase):
    """SortableAdminMixin (apps/core/admin.py): arrastrar filas en el
    listado del admin en vez de editar el número de `order` a mano. Se
    prueba sobre ContactLink por ser el modelo más simple de los que lo
    usan (Theme, Product, OscarCategory, PersonalityCharacter/Question)."""

    def setUp(self):
        self.admin = User.objects.create(email="drag_admin@test.local", role=User.Role.ADMIN, is_staff=True, is_superuser=True)
        self.admin.set_password("Testpass123!")
        self.admin.save()
        self.client.login(username=self.admin.email, password="Testpass123!")
        self.a = ContactLink.objects.create(platform=ContactLink.Platform.INSTAGRAM, label="a", url="https://a.example.com", order=0)
        self.b = ContactLink.objects.create(platform=ContactLink.Platform.WHATSAPP, label="b", url="https://b.example.com", order=1)
        self.c = ContactLink.objects.create(platform=ContactLink.Platform.TWITTER, label="c", url="https://c.example.com", order=2)

    def test_arrastrar_actualiza_el_orden_de_todos(self):
        url = reverse("admin:core_contactlink_reorder")
        response = self.client.post(
            url, data='{"order": [%d, %d, %d]}' % (self.c.pk, self.a.pk, self.b.pk),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.a.refresh_from_db()
        self.b.refresh_from_db()
        self.c.refresh_from_db()
        self.assertEqual(self.c.order, 0)
        self.assertEqual(self.a.order, 1)
        self.assertEqual(self.b.order, 2)

    def test_el_numero_ya_no_sale_en_el_formulario(self):
        response = self.client.get(reverse("admin:core_contactlink_change", args=[self.a.pk]))
        self.assertNotContains(response, 'name="order"')

    def test_el_listado_tiene_el_tirador_de_arrastre(self):
        response = self.client.get(reverse("admin:core_contactlink_changelist"))
        self.assertContains(response, "drag-handle")

    def test_requiere_estar_conectado_como_admin(self):
        self.client.logout()
        url = reverse("admin:core_contactlink_reorder")
        response = self.client.post(url, data="{}", content_type="application/json")
        self.assertNotEqual(response.status_code, 200)

    def test_no_admite_get(self):
        url = reverse("admin:core_contactlink_reorder")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 405)


class IntroLightThemeTests(TestCase):
    def test_theme_css_incluye_los_colores_de_la_animacion_de_inicio(self):
        theme = Theme.objects.get(slug="cinephile")
        self.client.post(reverse("core:set-theme", args=[theme.slug]))
        response = self.client.get(reverse("theme-css"))
        self.assertContains(response, "--color-intro-light")
        self.assertContains(response, "--color-intro-lamp")
        self.assertContains(response, "--color-intro-chair")
        self.assertContains(response, theme.color_intro_light)


@override_settings(VAPID_PUBLIC_KEY="clave-publica", VAPID_PRIVATE_KEY="clave-privada")
class PushHelperTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(email="push_helper@test.local", role=User.Role.LECTOR)
        self.subscription = PushSubscription.objects.create(
            user=self.user, endpoint="https://push.example/uno", p256dh="p", auth="a",
        )

    def test_push_enabled_requiere_claves_vapid(self):
        self.assertTrue(push_module.push_enabled())

    @override_settings(VAPID_PUBLIC_KEY="", VAPID_PRIVATE_KEY="")
    def test_push_disabled_sin_claves(self):
        self.assertFalse(push_module.push_enabled())

    @override_settings(VAPID_PUBLIC_KEY="", VAPID_PRIVATE_KEY="")
    @patch("apps.core.push.webpush")
    def test_no_manda_nada_si_esta_desactivado(self, mock_webpush):
        push_module.send_push_to_user(self.user, "Título", "Cuerpo")
        mock_webpush.assert_not_called()

    @patch("apps.core.push.webpush")
    def test_manda_push_a_cada_suscripcion_del_usuario(self, mock_webpush):
        push_module.send_push_to_user(self.user, "Título", "Cuerpo", url="/algo/")
        mock_webpush.assert_called_once()
        kwargs = mock_webpush.call_args.kwargs
        self.assertEqual(kwargs["subscription_info"]["endpoint"], "https://push.example/uno")

    @patch("apps.core.push.webpush")
    def test_borra_la_suscripcion_si_el_endpoint_ya_no_es_valido(self, mock_webpush):
        response = type("Response", (), {"status_code": 410})()
        mock_webpush.side_effect = push_module.WebPushException("caducada", response=response)
        push_module.send_push_to_user(self.user, "Título", "Cuerpo")
        self.assertFalse(PushSubscription.objects.filter(pk=self.subscription.pk).exists())

    @patch("apps.core.push.webpush")
    def test_conserva_la_suscripcion_si_falla_por_otro_motivo(self, mock_webpush):
        response = type("Response", (), {"status_code": 500})()
        mock_webpush.side_effect = push_module.WebPushException("error servidor", response=response)
        push_module.send_push_to_user(self.user, "Título", "Cuerpo")
        self.assertTrue(PushSubscription.objects.filter(pk=self.subscription.pk).exists())


@override_settings(GOOGLE_OAUTH_CLIENT_ID="client-id", GOOGLE_OAUTH_CLIENT_SECRET="client-secret")
class GoogleCalendarServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(email="gcal_service@test.local", role=User.Role.LECTOR)
        self.connection = GoogleCalendarConnection.objects.create(
            user=self.user, refresh_token="r", access_token="viejo",
            access_token_expires_at=timezone.now() - timezone.timedelta(hours=1),
        )

    def test_enabled_requiere_las_dos_credenciales(self):
        self.assertTrue(google_calendar_module.google_calendar_enabled())

    @override_settings(GOOGLE_OAUTH_CLIENT_ID="", GOOGLE_OAUTH_CLIENT_SECRET="")
    def test_disabled_sin_credenciales(self):
        self.assertFalse(google_calendar_module.google_calendar_enabled())

    def test_get_authorization_url_incluye_el_client_id_y_el_state(self):
        url = google_calendar_module.get_authorization_url("https://example.com/callback", "el-state")
        self.assertIn("client_id=client-id", url)
        self.assertIn("state=el-state", url)
        self.assertIn("access_type=offline", url)

    @patch("apps.core.google_calendar.requests.post")
    def test_create_event_renueva_el_token_caducado_antes_de_llamar(self, mock_post):
        token_response = Mock(status_code=200)
        token_response.json.return_value = {"access_token": "nuevo", "expires_in": 3600}
        event_response = Mock(status_code=200)
        event_response.json.return_value = {"id": "google-event-id"}
        mock_post.side_effect = [token_response, event_response]

        event_id = google_calendar_module.create_event(self.connection, "Título", date(2026, 3, 15))

        self.assertEqual(event_id, "google-event-id")
        self.connection.refresh_from_db()
        self.assertEqual(self.connection.access_token, "nuevo")
        # Primera llamada: refrescar el token. Segunda: crear el evento con
        # el token ya renovado en la cabecera Authorization.
        self.assertEqual(mock_post.call_args_list[1].kwargs["headers"]["Authorization"], "Bearer nuevo")

    @patch("apps.core.google_calendar.requests.post")
    def test_create_event_reutiliza_el_token_si_no_ha_caducado(self, mock_post):
        self.connection.access_token = "vigente"
        self.connection.access_token_expires_at = timezone.now() + timezone.timedelta(hours=1)
        self.connection.save()

        event_response = Mock(status_code=200)
        event_response.json.return_value = {"id": "e2"}
        mock_post.return_value = event_response

        google_calendar_module.create_event(self.connection, "Título", date(2026, 3, 15))

        mock_post.assert_called_once()
        self.assertEqual(mock_post.call_args.kwargs["headers"]["Authorization"], "Bearer vigente")

    @patch("apps.core.google_calendar.requests.delete")
    def test_delete_event_ignora_un_404_ya_borrado(self, mock_delete):
        self.connection.access_token = "vigente"
        self.connection.access_token_expires_at = timezone.now() + timezone.timedelta(hours=1)
        self.connection.save()
        mock_delete.return_value = Mock(status_code=404)

        google_calendar_module.delete_event(self.connection, "id-inexistente")  # no debe lanzar excepción


class TemplateCommentSyntaxTests(TestCase):
    """Django no soporta comentarios {# ... #} de varias líneas — si {# no
    se cierra con #} en la MISMA línea, el texto se filtra tal cual en la
    página en vez de desaparecer (nos pasó dos veces ya: templates/admin/
    actions.html y templates/base.html). {% comment %}...{% endcomment %}
    sí puede abarcar varias líneas — ese no es el problema aquí."""

    def test_no_hay_comentarios_llave_almohadilla_sin_cerrar_en_la_misma_linea(self):
        base_dir = Path(settings.BASE_DIR) / "templates"
        offenders = []
        for path in base_dir.rglob("*.html"):
            for lineno, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
                start = 0
                while True:
                    idx = line.find("{#", start)
                    if idx == -1:
                        break
                    if "#}" not in line[idx:]:
                        offenders.append(f"{path.relative_to(base_dir)}:{lineno}")
                    start = idx + 2
        self.assertEqual(offenders, [], f"Comentarios {{# #}} sin cerrar en la misma línea: {offenders}")


def _make_user(email, role=None):
    from apps.accounts.models import User as UserModel

    user = UserModel(email=email, role=role or UserModel.Role.LECTOR)
    user.set_password("Testpass123!")
    user.save()
    return user


class NotificationsBellTests(TestCase):
    """Campanita de la cabecera: agrega mensajes sin leer, solicitudes de
    amistad, avisos del equipo (Announcement) y artículos nuevos que no has
    visto — sin duplicar datos, cada cosa se sigue marcando como leída donde
    ya se marcaba (menos Announcement, que es nuevo y se marca al abrir)."""

    def setUp(self):
        self.user = _make_user("campana@test.local")
        self.other = _make_user("otra_persona@test.local")
        self.client.login(username=self.user.email, password="Testpass123!")

    def test_sin_nada_pendiente_el_contador_es_cero(self):
        self.assertEqual(unread_notifications_count(self.user), 0)

    def test_un_mensaje_sin_leer_cuenta(self):
        from apps.social.models import Message

        Message.objects.create(sender=self.other, recipient=self.user, body="Hola")
        self.assertEqual(unread_notifications_count(self.user), 1)
        feed = notifications_feed(self.user)
        self.assertEqual(feed[0]["kind"], "message")

    def test_una_solicitud_de_amistad_pendiente_cuenta(self):
        from apps.social.models import FriendRequest

        FriendRequest.objects.create(from_user=self.other, to_user=self.user)
        self.assertEqual(unread_notifications_count(self.user), 1)

    def test_un_articulo_nuevo_que_no_has_visto_cuenta(self):
        Article.objects.create(title="Recién publicado", body="<p>x</p>", author=self.other)
        self.assertEqual(unread_notifications_count(self.user), 1)

    def test_tu_propio_articulo_no_cuenta_como_pendiente(self):
        Article.objects.create(title="Mío", body="<p>x</p>", author=self.user)
        self.assertEqual(unread_notifications_count(self.user), 0)

    def test_un_articulo_ya_visto_deja_de_contar(self):
        from apps.articles.models import ArticleView

        article = Article.objects.create(title="Visto", body="<p>x</p>", author=self.other)
        ArticleView.objects.create(article=article, user=self.user)
        self.assertEqual(unread_notifications_count(self.user), 0)

    def test_un_aviso_del_equipo_cuenta_hasta_que_se_abre_el_panel(self):
        Announcement.objects.create(title="Mantenimiento", body="El sábado a las 10.")
        self.assertEqual(unread_notifications_count(self.user), 1)

        response = self.client.get(reverse("core:notifications-panel"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mantenimiento")

        self.assertEqual(unread_notifications_count(self.user), 0)

    def test_el_panel_de_avisos_requiere_login(self):
        self.client.logout()
        response = self.client.get(reverse("core:notifications-panel"))
        self.assertNotEqual(response.status_code, 200)

    def test_un_articulo_nuevo_en_la_tienda_cuenta(self):
        from apps.shop.models import Product

        Product.objects.create(name="Taza de la Sala")
        self.assertEqual(unread_notifications_count(self.user), 1)
        feed = notifications_feed(self.user)
        self.assertEqual(feed[0]["kind"], "product")

    def test_desactivada_desde_admin_el_contador_es_cero(self):
        from apps.social.models import Message

        Message.objects.create(sender=self.other, recipient=self.user, body="Hola")
        self.assertEqual(unread_notifications_count(self.user), 1)

        config = SiteConfig.load()
        config.notifications_bell_enabled = False
        config.save()

        self.assertEqual(unread_notifications_count(self.user), 0)
        self.assertEqual(notifications_feed(self.user), [])

    def test_desactivada_desde_admin_no_sale_la_campanita_en_la_cabecera(self):
        config = SiteConfig.load()
        config.notifications_bell_enabled = False
        config.save()

        response = self.client.get(reverse("core:home"))
        self.assertNotContains(response, "notif-bell")

    def test_un_aviso_del_equipo_sale_en_el_admin_de_usuarios_al_crearlo(self):
        admin = _make_user("admin_avisos@test.local", role=User.Role.ADMIN)
        self.client.login(username=admin.email, password="Testpass123!")
        self.client.post(reverse("admin:core_announcement_add"), {
            "title": "Aviso de prueba", "body": "Cuerpo", "url": "",
        })
        announcement = Announcement.objects.get(title="Aviso de prueba")
        self.assertEqual(announcement.created_by, admin)


class ExportExcelTests(TestCase):
    """Copia de seguridad en Excel desde el admin (enlace en el topmenu de
    Jazzmin): calendario, Top Secret, guardadas por lista y tienda, cada uno
    en su propia hoja."""

    def setUp(self):
        self.admin = _make_user("export_admin@test.local", role=User.Role.ADMIN)
        self.admin.is_staff = True
        self.admin.is_superuser = True
        self.admin.save(update_fields=["is_staff", "is_superuser"])
        self.client.login(username=self.admin.email, password="Testpass123!")

    def test_requiere_ser_staff(self):
        self.client.logout()
        lector = _make_user("export_lector@test.local")
        self.client.login(username=lector.email, password="Testpass123!")
        response = self.client.get(reverse("core:admin-export-excel"))
        self.assertNotEqual(response.status_code, 200)

    def test_devuelve_un_excel_descargable(self):
        response = self.client.get(reverse("core:admin-export-excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

    def test_incluye_las_cuatro_hojas_con_sus_datos(self):
        import io

        from openpyxl import load_workbook

        from apps.movies.models import Movie, SavedMovie, SavedMovieList
        from apps.secret.models import CalendarDayNote, Genre, SecretMovie
        from apps.shop.models import Product

        user = _make_user("export_datos@test.local")
        CalendarDayNote.objects.create(user=user, date=date(2026, 1, 5), note="Maratón de Navidad")

        terror = Genre.objects.create(name="Terror")
        movie_secret = SecretMovie.objects.create(title="El Exorcista", personal_rating="9.5", comment="Un clásico.")
        movie_secret.genres.add(terror)

        movie = Movie.objects.create(tmdb_id=1, title="Drive", year="2011")
        sublist = SavedMovieList.objects.create(user=user, name="Favoritas")
        saved = SavedMovie.objects.create(user=user, movie=movie)
        saved.sublists.add(sublist)

        Product.objects.create(name="Taza Bygui", description="Con el logo.", price="12.50", url="https://example.com/taza")

        response = self.client.get(reverse("core:admin-export-excel"))
        wb = load_workbook(io.BytesIO(response.content))

        self.assertEqual(wb.sheetnames, ["Calendario", "Top Secret", "Guardadas", "Tienda"])

        calendar_rows = list(wb["Calendario"].iter_rows(values_only=True))
        self.assertIn((str(user), "05/01/2026", "Maratón de Navidad"), calendar_rows)

        secret_rows = list(wb["Top Secret"].iter_rows(values_only=True))
        self.assertIn(("El Exorcista", 9.5, "Terror", "Un clásico."), secret_rows)

        saved_rows = list(wb["Guardadas"].iter_rows(values_only=True))
        self.assertIn((str(user), "Favoritas", "Drive"), saved_rows)

        shop_rows = list(wb["Tienda"].iter_rows(values_only=True))
        self.assertIn(("Taza Bygui", "Con el logo.", 12.5, "https://example.com/taza"), shop_rows)
