from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.response import TemplateResponse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_POST

from apps.articles.models import Article
from apps.articles.permissions import can_manage_private_articles

from .forms import ContactForm
from .models import SESSION_THEME_KEY, Announcement, ContactLink, SiteConfig, Theme, get_effective_theme
from .notifications import notifications_feed, unread_notifications_count


def home(request):
    from .activity import recent_activity

    articles = Article.objects.select_related("author").prefetch_related("tags")
    if not can_manage_private_articles(request.user):
        articles = articles.filter(is_private=False)

    # Si hay artículos marcados a mano como destacados, esos son los que
    # rotan en el carrusel; si no se ha marcado ninguno, se usan los más
    # recientes (comportamiento de siempre). En ambos casos, "Últimos
    # artículos" no repite ninguno de los que ya salen en el carrusel.
    featured = list(articles.filter(is_featured=True)[:5])
    if featured:
        hero_articles = featured[:3]
        grid_articles = list(articles.exclude(pk__in=[a.pk for a in featured])[:5])
    else:
        recent = list(articles[:8])
        hero_articles = recent[:3]
        grid_articles = recent[3:8]

    return TemplateResponse(request, "core/home.html", {
        "hero_articles": hero_articles,
        "grid_articles": grid_articles,
        "recent_activity": recent_activity(request.user),
    })


@login_required
def notifications_panel(request):
    """Contenido de la campanita — se pide por HTMX al abrir el
    desplegable. Abrirlo ya marca como leídos los avisos del equipo
    (Announcement, lo único que gestiona esta campanita de verdad); los
    mensajes/solicitudes/artículos se marcan leídos donde ya se marcaban
    antes (su propia página), esto solo enseña que los tienes pendientes.

    El número que queda tras abrir (mensajes/solicitudes/artículos/
    productos sin visitar todavía, que abrir la campanita no marca como
    vistos) va en una cabecera aparte: antes el JS quitaba el globo entero
    sin más al abrir, así que en la siguiente recarga volvía a aparecer con
    el mismo número — parecía que "no se acordaba" de que ya lo habías
    abierto."""
    feed = notifications_feed(request.user)
    for announcement in Announcement.objects.exclude(read_by=request.user):
        announcement.read_by.add(request.user)
    response = render(request, "core/_notifications_panel.html", {"feed": feed})
    response["X-Notif-Remaining"] = str(unread_notifications_count(request.user))
    return response


def donations(request):
    return render(request, "core/donations.html", {"site_config": SiteConfig.load()})


def contact(request):
    contact_links = ContactLink.objects.all()

    if request.method == "POST":
        form = ContactForm(request.POST)
        if form.is_valid():
            if not form.is_spam():
                _notify_admins_of_contact_message(form.cleaned_data["name"], form.cleaned_data["email"], form.cleaned_data["message"])
            messages.success(request, "¡Mensaje enviado! Te responderemos lo antes posible.")
            return redirect("core:contact")
    else:
        form = ContactForm()

    return render(request, "core/contact.html", {"form": form, "contact_links": contact_links})


def _notify_admins_of_contact_message(name, email, message):
    """"Escríbenos" ya no manda un email (el SMTP no era de fiar) — en su
    lugar, cada Admin recibe el mensaje como un aviso de Social (mensaje
    autoenviado, ya que quien escribe no tiene por qué tener cuenta), donde
    también aparece en su campanita de avisos."""
    from apps.accounts.models import User
    from apps.social.models import Message

    body = f"Mensaje de «Escríbenos» de {name} <{email}>:\n\n{message}"
    for admin in User.objects.filter(role=User.Role.ADMIN):
        Message.objects.create(sender=admin, recipient=admin, body=body)


@cache_control(private=True, no_cache=True)
def theme_css(request):
    theme = get_effective_theme(request.user, request.session)
    return TemplateResponse(
        request, "core/theme.css", {"theme": theme}, content_type="text/css"
    )


def _theme_redirect_or_json(request, slug):
    """El selector de temas de la cabecera lo llama por fetch() (espera
    JSON), pero el de Ajustes usa un <form> normal con un campo oculto
    "next" — sin JS que reaccione a la respuesta, así que ahí hace falta
    una redirección de verdad en vez de un JSON en pantalla."""
    next_url = request.POST.get("next")
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        return redirect(next_url)
    return JsonResponse({"ok": True, "slug": slug})


@require_POST
def set_theme(request, slug):
    theme = get_object_or_404(Theme, slug=slug)
    if request.user.is_authenticated:
        request.user.theme = theme
        request.user.save(update_fields=["theme"])
    else:
        request.session[SESSION_THEME_KEY] = theme.slug
    return _theme_redirect_or_json(request, theme.slug)


@require_POST
def reset_theme(request):
    if request.user.is_authenticated:
        request.user.theme = None
        request.user.save(update_fields=["theme"])
    request.session.pop(SESSION_THEME_KEY, None)
    return _theme_redirect_or_json(request, None)


def service_worker(request):
    """Se sirve en /sw.js (no en /static/js/sw.js): el scope por defecto de
    un service worker es el directorio de su propia URL, así que si viviera
    bajo /static/js/ nunca podría controlar el resto del sitio."""
    path = finders.find("js/sw.js")
    if not path:
        raise Http404
    response = FileResponse(open(path, "rb"), content_type="application/javascript")
    response["Service-Worker-Allowed"] = "/"
    return response


@staff_member_required
def export_excel(request):
    """Copia de seguridad rápida en Excel de lo que no está en ningún otro
    sitio si la web "explota": el comentario de cada día del calendario
    personal y las películas/series que tenga puestas ese día, la lista
    de Top Secret (nota, listas, comentario, estado de visionado de las
    series), las descripciones del tablón de fotos de cada quien (las
    imágenes en sí no, solo el texto — para eso está el propio storage),
    las guardadas de cada usuario agrupadas por su lista, el catálogo de
    la Tienda, los imprescindibles/sugeridas de cada perfil con su
    porqué, las ideas de artículos, las frases favoritas y los apuntes
    personales del admin — en resumen, todo lo que alguien ha escrito a
    mano y no está en ningún otro sitio. Un botón en el panel (topmenu
    de Jazzmin), nada que requiera entrar a ningún modelo."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from apps.accounts.models import FavoriteMovie
    from apps.articles.models import ArticleIdea
    from apps.movies.models import SavedMovie
    from apps.secret.models import CalendarDayNote, ReleaseEvent, SecretMovie, SecretPhoto
    from apps.shop.models import Product
    from .models import FavoriteQuote, PersonalNote

    wb = Workbook()

    ws_calendar = wb.active
    ws_calendar.title = "Calendario"
    ws_calendar.append(["Usuario", "Fecha", "Comentario"])
    for note in CalendarDayNote.objects.select_related("user").order_by("user__username", "date"):
        ws_calendar.append([str(note.user), note.date.strftime("%d/%m/%Y"), note.note])

    ws_calendar_movies = wb.create_sheet("Calendario - películas")
    ws_calendar_movies.append(["Usuario", "Fecha", "Película/Serie", "Nota"])
    for event in ReleaseEvent.objects.select_related("user", "movie").order_by("user__username", "date"):
        ws_calendar_movies.append([str(event.user), event.date.strftime("%d/%m/%Y"), event.movie.title, event.note])

    ws_secret = wb.create_sheet("Top Secret")
    ws_secret.append(["Nombre", "Nota", "Listas", "Comentario", "Estado (series)"])
    for movie in SecretMovie.objects.prefetch_related("genres").select_related("movie").order_by("title"):
        listas = ", ".join(movie.genres.values_list("name", flat=True))
        estado = movie.get_series_watch_status_display() if movie.movie_id and movie.movie.is_tv else ""
        ws_secret.append([movie.title, float(movie.personal_rating), listas, movie.comment, estado])

    ws_saved = wb.create_sheet("Guardadas")
    ws_saved.append(["Usuario", "Lista", "Película"])
    for saved in SavedMovie.objects.select_related("user", "movie").prefetch_related("sublists").order_by("user__username", "movie__title"):
        listas = ", ".join(saved.sublists.values_list("name", flat=True)) or "Sin lista"
        ws_saved.append([str(saved.user), listas, saved.movie.title])

    ws_photo_board = wb.create_sheet("Tablón de fotos")
    ws_photo_board.append(["Tablón de", "Subida por", "Descripción", "Fecha"])
    for photo in SecretPhoto.objects.select_related("board_owner", "uploaded_by").order_by("board_owner__username", "created_at"):
        ws_photo_board.append([
            str(photo.board_owner), str(photo.uploaded_by), photo.description,
            photo.created_at.strftime("%d/%m/%Y %H:%M"),
        ])

    ws_shop = wb.create_sheet("Tienda")
    ws_shop.append(["Nombre", "Descripción", "Precio", "Enlace"])
    for product in Product.objects.order_by("order", "name"):
        ws_shop.append([product.name, product.description, float(product.price) if product.price is not None else None, product.url])

    ws_favorites = wb.create_sheet("Imprescindibles y sugeridas")
    ws_favorites.append(["Usuario", "Categoría", "Película", "Por qué (de toda la categoría)"])
    notes_by_user = {}
    for fav in FavoriteMovie.objects.select_related("user", "movie").order_by("user__username", "category", "order"):
        key = (fav.user_id, fav.category)
        if key not in notes_by_user:
            note_field = "essential_note" if fav.category == FavoriteMovie.Category.ESSENTIAL else "suggested_note"
            notes_by_user[key] = getattr(fav.user, note_field)
        ws_favorites.append([str(fav.user), fav.get_category_display(), fav.movie.title, notes_by_user[key]])

    ws_ideas = wb.create_sheet("Ideas de artículos")
    ws_ideas.append(["Idea", "Notas", "Ya escrito", "Apuntada por", "Fecha"])
    for idea in ArticleIdea.objects.select_related("created_by").order_by("-created_at"):
        ws_ideas.append([
            idea.text, idea.notes, "Sí" if idea.is_done else "No",
            str(idea.created_by) if idea.created_by else "", idea.created_at.strftime("%d/%m/%Y"),
        ])

    ws_quotes = wb.create_sheet("Frases favoritas")
    ws_quotes.append(["Frase", "Película o serie", "Notas", "Fecha"])
    for quote in FavoriteQuote.objects.order_by("-created_at"):
        ws_quotes.append([quote.text, quote.source, quote.notes, quote.created_at.strftime("%d/%m/%Y")])

    ws_notes = wb.create_sheet("Apuntes personales")
    ws_notes.append(["Título", "Apunte", "Creado", "Última edición"])
    for note in PersonalNote.objects.order_by("-updated_at"):
        ws_notes.append([note.title, note.body, note.created_at.strftime("%d/%m/%Y"), note.updated_at.strftime("%d/%m/%Y")])

    for ws in wb.worksheets:
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"Lasaladebyguibackup_{timezone.now():%Y%m%d_%H%M}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
